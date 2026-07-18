"""Report generation for validation outputs."""

from __future__ import annotations

from dataclasses import asdict
from pathlib import Path

import pandas as pd
from openpyxl.styles import PatternFill

from .excel_processor import ExtractedTable
from .validator import ValidationRecord, validate_totals


RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")


def build_validation_results(
    tables: list[ExtractedTable], tolerance: float = 0.01
) -> tuple[list[tuple[ExtractedTable, pd.DataFrame]], pd.DataFrame]:
    details: list[tuple[ExtractedTable, pd.DataFrame]] = []
    records: list[ValidationRecord] = []

    for table in tables:
        validated_df, table_records = validate_totals(table.dataframe, table.table_name, tolerance=tolerance)
        details.append((table, validated_df))
        records.extend(table_records)

    summary_df = pd.DataFrame([asdict(r) for r in records])
    if summary_df.empty:
        summary_df = pd.DataFrame(
            columns=[
                "table_name",
                "column_name",
                "reported_total",
                "computed_total",
                "difference",
                "status",
                "row_index",
            ]
        )

    summary_df = summary_df.rename(
        columns={
            "table_name": "Table name / Sheet name",
            "column_name": "Column name",
            "reported_total": "Reported total",
            "computed_total": "Computed total",
            "difference": "Difference",
            "status": "Status",
        }
    )

    return details, summary_df


def generate_excel_report(
    output_path: str,
    tables: list[ExtractedTable],
    tolerance: float = 0.01,
) -> str:
    """Generate final Excel with summary + detailed sheets + highlighting."""
    details, summary_df = build_validation_results(tables, tolerance=tolerance)

    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="Summary Report", index=False)

        for idx, (table, validated_df) in enumerate(details, start=1):
            sheet_name = f"Detail_{idx}"
            validated_df.to_excel(writer, sheet_name=sheet_name, index=False)

        workbook = writer.book
        summary_ws = workbook["Summary Report"]
        status_col = None
        for cidx, cell in enumerate(summary_ws[1], start=1):
            if cell.value == "Status":
                status_col = cidx
                break

        if status_col is not None:
            for row in range(2, summary_ws.max_row + 1):
                status = summary_ws.cell(row=row, column=status_col).value
                if status == "Casting Error":
                    for col in range(1, summary_ws.max_column + 1):
                        summary_ws.cell(row=row, column=col).fill = RED_FILL
                elif status == "Decimal Rounding Difference":
                    for col in range(1, summary_ws.max_column + 1):
                        summary_ws.cell(row=row, column=col).fill = YELLOW_FILL

        for idx in range(1, len(details) + 1):
            ws = workbook[f"Detail_{idx}"]
            headers = [ws.cell(row=1, column=i).value for i in range(1, ws.max_column + 1)]
            try:
                status_col_idx = headers.index("status") + 1
            except ValueError:
                continue
            for row in range(2, ws.max_row + 1):
                status = ws.cell(row=row, column=status_col_idx).value
                if status == "Casting Error":
                    for col in range(1, ws.max_column + 1):
                        ws.cell(row=row, column=col).fill = RED_FILL
                elif status == "Decimal Rounding Difference":
                    for col in range(1, ws.max_column + 1):
                        ws.cell(row=row, column=col).fill = YELLOW_FILL

    return str(output)
